
#! RUN THIS FILE TO GENERATE THE ACTUAL TIMETABLE EXCEL FILE    

from __future__ import annotations
from pathlib import Path
import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# -------------------------------
# DATA (change with user data)
# -------------------------------

classes = ["6A", "6B"]
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
periods = [1, 2, 3, 4, 5, 6]
#! Change the subjects and their respective teachers and hours as per your requirements
#! Total should be 30 hours per week (6 periods * 5 days)
subjects = {
    "Math": {"teacher": "Mr. Sharma", "hours": 6},
    "English": {"teacher": "Ms. Sen", "hours": 6},
    "Science": {"teacher": "Mr. Das", "hours": 5},
    "History": {"teacher": "Mr. Roy", "hours": 4},
    "Computer": {"teacher": "Mrs. Gupta", "hours": 4},
    "Free": {"teacher": "FREE", "hours": 5},
}

#-------------------------------
#EXCEL FILE GENERATION
#-------------------------------
def save_timetable_excel(rows,
                         filename="generated_timetable.xlsx",
                         school_name="ABC SCHOOL"):

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="4472C4")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    cell_fill = PatternFill("solid", fgColor="DCE6F1")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=16)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    all_classes = sorted({r[0] for r in rows})

    for cls in all_classes:

        ws = wb.create_sheet(title=str(cls))

        # Title
        ws.merge_cells(start_row=1,
                       start_column=1,
                       end_row=1,
                       end_column=len(days) + 1)

        title = ws["A1"]
        title.value = f"{school_name} - Class {cls} Timetable"
        title.fill = title_fill
        title.font = title_font
        title.alignment = center

        # Headers
        ws["A2"] = "Period"

        ws["A2"].fill = header_fill
        ws["A2"].font = white_font
        ws["A2"].alignment = center

        for col, day in enumerate(days, start=2):

            cell = ws.cell(row=2, column=col)
            cell.value = day
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        # Period labels
        for row_no, period in enumerate(periods, start=3):

            cell = ws.cell(row=row_no, column=1)
            cell.value = f"Period {period}"
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        # Fill timetable
        for cls_name, day, period, subject, teacher in rows:

            if cls_name != cls:
                continue

            row = periods.index(period) + 3
            col = days.index(day) + 2

            cell = ws.cell(row=row, column=col)
            cell.value = f"{subject}\n({teacher})"
            cell.alignment = center
            cell.fill = cell_fill
            cell.border = border

        ws.freeze_panes = "B3"

        # Column widths
        ws.column_dimensions["A"].width = 14

        for c in range(2, len(days) + 2):
            ws.column_dimensions[get_column_letter(c)].width = 24

        # Row heights
        for r in range(3, len(periods) + 3):
            ws.row_dimensions[r].height = 45

    wb.save(filename)

    print(f"Excel timetable saved as {filename}")


#-------------------------------
# TIMETABLE GENERATION
#-------------------------------

def generate_timetable(output_path: str | None = None) -> pd.DataFrame:
    """Create a simple feasible timetable and save it to a CSV file."""
    model = cp_model.CpModel()

    # x[(class, day, period, subject)] = 1 if the subject is scheduled at that slot
    x = {}
    for cls in classes:
        for day in days:
            for period in periods:
                for subject in subjects:
                    x[(cls, day, period, subject)] = model.NewBoolVar(
                        f"{cls}_{day}_{period}_{subject}"
                    )

    # One subject per class, per day, per period
    for cls in classes:
        for day in days:
            for period in periods:
                model.Add(
                    sum(x[(cls, day, period, subject)] for subject in subjects) == 1
                )

    # Each subject must appear for the required number of hours for each class
    for cls in classes:
        for subject, info in subjects.items():
            model.Add(
                sum(
                    x[(cls, day, period, subject)]
                    for day in days
                    for period in periods
                )
                == info["hours"]
            )

    # Avoid assigning the same subject in consecutive periods on the same day
    for cls in classes:
        for day in days:
            for period in periods[:-1]:
                for subject in subjects:
                    model.Add(
                        x[(cls, day, period, subject)]
                        + x[(cls, day, period + 1, subject)]
                        <= 1
                    )

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"Unable to build a timetable. Solver status: {status}")

    rows = []
    for cls in classes:
        for day in days:
            for period in periods:
                for subject in subjects:
                    if solver.Value(x[(cls, day, period, subject)]):
                        teacher = subjects[subject]["teacher"]
                        rows.append([cls, day, period, subject, teacher])

    df = pd.DataFrame(
        rows,
        columns=["Class", "Day", "Period", "Subject", "Teacher"],
    )

    rows = df.values.tolist()
    output_file = Path(output_path or "generated_timetable.xlsx")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    save_timetable_excel(
        rows,
        filename=str(output_file),
        school_name="ABC SCHOOL"
    )
    return df


if __name__ == "__main__":
    timetable = generate_timetable()
    print("\nTimetable Generated Successfully\n")
    print(timetable.to_string(index=False))
    print("\nExcel timetable saved successfully!")